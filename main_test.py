"""
Fast, per-step manual verification script.
Run this after every implementation step with: python main_test.py
Full pytest suite is reserved for milestone boundaries only.
"""

from app.core.logging_setup import configure_logging
from loguru import logger


def test_config_loads():
    """M1-S2: confirm settings load and required paths resolve correctly."""
    from app.core.config import settings

    assert settings.data_dir_path.exists(), "data_dir_path should exist"
    assert settings.logs_dir_path.exists(), "logs_dir_path should exist"
    print(f"[M1-S2] Config OK — app_env={settings.app_env}, log_level={settings.log_level}")
    print(f"[M1-S2] db_path resolved to: {settings.db_path_resolved}")


def test_logging_writes():
    """M1-S2: confirm logger writes to console and log file without error."""
    logger.debug("Debug-level test message")
    logger.info("Info-level test message")
    logger.warning("Warning-level test message")
    print("[M1-S2] Logging OK — check logs/agent.log for file output")

def test_database_init():
    """M1-S3: confirm init_db() runs cleanly and execution_history table exists."""
    from app.core.database import init_db, connection

    init_db()

    with connection() as conn:
        cursor = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='execution_history'"
        )
        row = cursor.fetchone()
        assert row is not None, "execution_history table should exist after init_db()"

    print("[M1-S3] Database OK - execution_history table confirmed")

def test_plugin_discovery():
    """M1-S4: confirm recursive discovery finds both the top-level AND
    the nested example tool - proves walk_packages actually recurses."""
    from app.registry.discovery import discover_tools
    from app.registry.tool_contract import get_registry

    count = discover_tools()
    registry = get_registry()

    assert "example_ping" in registry, "top-level example tool should be registered"
    assert "example_deep_ping" in registry, "nested example tool should be registered - if missing, recursion is broken"
    print(f"[M1-S4] Registry OK - {count} tool(s) registered: {list(registry.keys())}")

def test_call_tool_success():
    """M2-S1: confirm call_tool successfully invokes a registered tool
    and returns a populated ToolResult - never None."""
    from app.core.call_tool import call_tool

    result = call_tool("example_ping", {"message": "hello M2"})
    assert result is not None, "call_tool must never return None"
    assert result.success is True
    assert result.data == "hello M2"
    print(f"[M2-S1] call_tool success path OK - result={result.data}")


def test_call_tool_unknown_tool():
    """M2-S1: confirm call_tool raises ToolExecutionError for an unregistered tool name."""
    from app.core.call_tool import call_tool
    from app.core.exceptions import ToolExecutionError

    try:
        call_tool("this_tool_does_not_exist", {})
        assert False, "call_tool should have raised for an unknown tool"
    except ToolExecutionError:
        print("[M2-S1] call_tool unknown-tool path OK - raised ToolExecutionError as expected")


def test_call_tool_invalid_input():
    """M2-S1: confirm invalid input is rejected BEFORE the tool function runs
    - directly tests the 'missing input_schema silently allows bad input' lesson."""
    from app.core.call_tool import call_tool
    from app.core.exceptions import ValidationError

    try:
        # example_ping expects `message: str` - passing an int should fail Pydantic validation
        call_tool("example_ping", {"message": {"not": "a string or int-coercible value"}})
        assert False, "call_tool should have raised ValidationError for bad input"
    except ValidationError:
        print("[M2-S1] call_tool input-validation path OK - raised ValidationError as expected")


def test_call_tool_audit_logged():
    """M2-S1: confirm every call_tool invocation writes a row to execution_history,
    regardless of success/failure - the Section 2 'full audit trail' requirement."""
    from app.core.call_tool import call_tool
    from app.core.database import connection

    with connection() as conn:
        before = conn.execute("SELECT COUNT(*) as c FROM execution_history").fetchone()["c"]

    call_tool("example_ping", {"message": "audit check"})

    with connection() as conn:
        after = conn.execute("SELECT COUNT(*) as c FROM execution_history").fetchone()["c"]

    assert after == before + 1, "exactly one execution_history row should be written per call_tool invocation"
    print(f"[M2-S1] Audit logging OK - execution_history grew from {before} to {after}")

def test_call_tool_read_needs_no_approval():
    """M2-S2: READ-permission tools should never require an approval_handler."""
    from app.core.call_tool import call_tool

    result = call_tool("example_ping", {"message": "no approval needed"})
    assert result.success is True
    print("[M2-S2] READ tool bypassed approval gate correctly")


def test_call_tool_modify_denied_without_handler():
    """M2-S2: MODIFY-permission tool with NO approval_handler must fail loudly,
    never silently skip the gate."""
    from app.core.call_tool import call_tool
    from app.core.exceptions import ToolExecutionError

    try:
        call_tool("example_modify", {"value": "x"})
        assert False, "should have raised - no approval_handler was provided for a MODIFY tool"
    except ToolExecutionError:
        print("[M2-S2] MODIFY tool correctly refused to run without an approval_handler")


def test_call_tool_modify_approved():
    """M2-S2: MODIFY tool runs when AutoApproveHandler approves it."""
    from app.core.call_tool import call_tool
    from app.core.approval import AutoApproveHandler

    result = call_tool("example_modify", {"value": "approved run"}, approval_handler=AutoApproveHandler())
    assert result.success is True
    print(f"[M2-S2] MODIFY tool approved and ran - result={result.data}")


def test_call_tool_modify_denied():
    """M2-S2: MODIFY tool raises when AutoDenyHandler denies it."""
    from app.core.call_tool import call_tool
    from app.core.approval import AutoDenyHandler
    from app.core.exceptions import ToolExecutionError

    try:
        call_tool("example_modify", {"value": "denied run"}, approval_handler=AutoDenyHandler())
        assert False, "should have raised - AutoDenyHandler denies everything"
    except ToolExecutionError:
        print("[M2-S2] MODIFY tool correctly blocked by denial")


def test_call_tool_batch_approval_per_run():
    """M2-S2: two calls sharing a run_id should only prompt/approve ONCE -
    the second call reuses the first's approval (Section 2 batch-approval rule)."""
    from app.core.call_tool import call_tool
    from app.core.approval import AutoDenyHandler  # denier - proves 2nd call did NOT re-prompt

    run_id = "test-batch-run-001"

    # First call: approve via AutoApproveHandler explicitly.
    from app.core.approval import AutoApproveHandler
    result1 = call_tool("example_modify", {"value": "first"}, run_id=run_id, approval_handler=AutoApproveHandler())
    assert result1.success is True

    # Second call, SAME run_id, but pass a DENIER this time - if batching
    # works, it should succeed anyway because the run was already approved
    # and the handler should never even be consulted.
    result2 = call_tool("example_modify", {"value": "second"}, run_id=run_id, approval_handler=AutoDenyHandler())
    assert result2.success is True, "second call in the same run_id should reuse the batch approval, not re-prompt"
    print("[M2-S2] Batch approval OK - second call in same run_id reused first call's approval")

def test_event_bus_basic_pubsub():
    """M3-S1: confirm subscribe/publish delivers events to a handler."""
    from app.core.event_bus import subscribe, publish, unsubscribe, clear_subscribers

    received = []

    def handler(event):
        received.append(event)

    clear_subscribers()
    subscribe("test.event", handler)
    publish("test.event", {"foo": "bar"})

    assert len(received) == 1, "handler should have received exactly one event"
    assert received[0].event_name == "test.event"
    assert received[0].payload == {"foo": "bar"}

    unsubscribe("test.event", handler)
    publish("test.event", {"foo": "baz"})
    assert len(received) == 1, "handler should not receive events after unsubscribing"

    print("[M3-S1] Event bus basic pub/sub OK")


def test_event_bus_subscriber_exception_does_not_propagate():
    """M3-S1: a broken subscriber must not crash the publisher."""
    from app.core.event_bus import subscribe, publish, clear_subscribers

    def broken_handler(event):
        raise RuntimeError("intentionally broken handler")

    clear_subscribers()
    subscribe("test.broken", broken_handler)

    # Should NOT raise, despite the subscriber blowing up internally.
    publish("test.broken", {})
    print("[M3-S1] Event bus subscriber-exception isolation OK")


def test_call_tool_publishes_events():
    """M3-S1: confirm call_tool publishes tool.succeeded and tool.failed
    at the right points, without those events firing for pre-execution
    rejections like an unknown tool name."""
    from app.core.event_bus import subscribe, clear_subscribers
    from app.core.call_tool import call_tool
    from app.core.exceptions import ToolExecutionError

    succeeded_events = []
    failed_events = []

    clear_subscribers()
    subscribe("tool.succeeded", lambda e: succeeded_events.append(e))
    subscribe("tool.failed", lambda e: failed_events.append(e))

    call_tool("example_ping", {"message": "event test"})
    assert len(succeeded_events) == 1
    assert succeeded_events[0].payload["tool_name"] == "example_ping"

    try:
        call_tool("nonexistent_tool", {})
    except ToolExecutionError:
        pass
    assert len(failed_events) == 0, "tool.failed should NOT fire for an unknown tool - it never ran"

    print("[M3-S1] call_tool event publishing OK - tool.succeeded fired, tool.failed correctly did not fire for unknown tool")

def test_approval_events():
    """M3-S2: confirm approval_requested/granted/denied fire at the right points."""
    from app.core.event_bus import subscribe, clear_subscribers
    from app.core.call_tool import call_tool
    from app.core.approval import AutoApproveHandler, AutoDenyHandler
    from app.core.exceptions import ToolExecutionError

    requested, granted, denied = [], [], []
    clear_subscribers()
    subscribe("tool.approval_requested", lambda e: requested.append(e))
    subscribe("tool.approval_granted", lambda e: granted.append(e))
    subscribe("tool.approval_denied", lambda e: denied.append(e))

    call_tool("example_modify", {"value": "x"}, approval_handler=AutoApproveHandler())
    assert len(requested) == 1 and len(granted) == 1 and len(denied) == 0

    try:
        call_tool("example_modify", {"value": "x"}, approval_handler=AutoDenyHandler())
    except ToolExecutionError:
        pass
    assert len(denied) == 1

    run_id = "main-test-batch"
    call_tool("example_modify", {"value": "x"}, run_id=run_id, approval_handler=AutoApproveHandler())
    call_tool("example_modify", {"value": "x"}, run_id=run_id, approval_handler=AutoDenyHandler())
    assert len(granted) == 3, "3rd and 4th calls: fresh + batch-reused, both grant"
    assert len(requested) == 3, "3 fresh calls request; only the batch-reused 4th call should NOT trigger a new request"
    print("[M3-S2] Approval events OK - requested/granted/denied fire at correct points")
def test_knowledge_base_crud():
    """M4-S1: confirm add/get/list/update/delete all work correctly."""
    from app.core.knowledge_base import (
        init_knowledge_base,
        add_knowledge_item,
        get_knowledge_item,
        list_knowledge_items,
        update_knowledge_item,
        delete_knowledge_item,
    )

    init_knowledge_base()

    item_id = add_knowledge_item("note", "Test note content", metadata={"source": "main_test"})
    assert item_id is not None

    item = get_knowledge_item(item_id)
    assert item["content"] == "Test note content"
    assert item["content_type"] == "note"
    assert item["metadata"] == {"source": "main_test"}
    assert item["embedding"] is None, "embedding should be nullable and unpopulated for now"

    items = list_knowledge_items(content_type="note")
    assert any(i["id"] == item_id for i in items)

    updated = update_knowledge_item(item_id, content="Updated content")
    assert updated is True
    assert get_knowledge_item(item_id)["content"] == "Updated content"

    deleted = delete_knowledge_item(item_id)
    assert deleted is True
    assert get_knowledge_item(item_id) is None

    print("[M4-S1] Knowledge base CRUD OK - add/get/list/update/delete all correct")


def test_knowledge_base_invalid_content_type():
    """M4-S1: confirm an invalid content_type is rejected, not silently accepted."""
    from app.core.knowledge_base import add_knowledge_item
    from app.core.exceptions import ValidationError

    try:
        add_knowledge_item("not_a_real_type", "content")
        assert False, "should have raised ValidationError for invalid content_type"
    except ValidationError:
        print("[M4-S1] Invalid content_type correctly rejected")
def test_knowledge_tools():
    """M4-S2: confirm the knowledge base @tool wrappers work end-to-end via call_tool."""
    from app.core.call_tool import call_tool
    from app.core.approval import AutoApproveHandler

    add_result = call_tool(
        "add_knowledge_item",
        {"content_type": "preference", "content": "prefers dark mode UIs"},
        approval_handler=AutoApproveHandler(),
    )
    assert add_result.success is True
    item_id = add_result.data["id"]

    search_result = call_tool("search_knowledge", {"query": "dark mode"})
    assert search_result.success is True
    assert any(i["id"] == item_id for i in search_result.data)

    delete_result = call_tool(
        "delete_knowledge_item", {"item_id": item_id}, approval_handler=AutoApproveHandler()
    )
    assert delete_result.success is True

    print("[M4-S2] Knowledge base tools OK - add/search/delete via call_tool all correct")
def test_notification_channel_console():
    """M5-S1: confirm ConsoleNotificationChannel sends successfully."""
    from app.core.notifications import ConsoleNotificationChannel

    channel = ConsoleNotificationChannel()
    result = channel.send("Test Title", "Test message body", metadata={"source": "main_test"})
    assert result is True
    print("[M5-S1] ConsoleNotificationChannel OK")


def test_notification_manager_broadcasts_to_all_channels():
    """M5-S1: confirm NotificationManager broadcasts to multiple channels
    and isolates a broken channel from the others."""
    from app.core.notifications import NotificationManager, NotificationChannel

    sent_log = []

    class WorkingChannel(NotificationChannel):
        @property
        def channel_name(self) -> str:
            return "working"

        def send(self, title, message, metadata=None) -> bool:
            sent_log.append((title, message))
            return True

    class BrokenChannel(NotificationChannel):
        @property
        def channel_name(self) -> str:
            return "broken"

        def send(self, title, message, metadata=None) -> bool:
            raise RuntimeError("intentionally broken channel")

    manager = NotificationManager()
    manager.register_channel(WorkingChannel())
    manager.register_channel(BrokenChannel())

    results = manager.notify("Broadcast Test", "hello all channels")

    assert results == {"working": True, "broken": False}
    assert len(sent_log) == 1
    print("[M5-S1] NotificationManager OK - broadcast to multiple channels, isolated the broken one")

def test_notification_event_wiring():
    """M5-S2: confirm approval_requested and tool.failed events trigger
    notifications, and tool.succeeded correctly does NOT."""
    from app.core.notifications import notification_manager, NotificationChannel
    from app.core.notification_wiring import wire_notifications_to_events
    from app.core.event_bus import publish

    received = []

    class TestChannel(NotificationChannel):
        @property
        def channel_name(self) -> str:
            return "test"

        def send(self, title, message, metadata=None) -> bool:
            received.append((title, message))
            return True

    channel = TestChannel()
    notification_manager.register_channel(channel)
    wire_notifications_to_events()

    publish("tool.approval_requested", {"tool_name": "x", "permission": "modify"})
    publish("tool.failed", {"tool_name": "y", "error": "boom"})
    publish("tool.succeeded", {"tool_name": "z", "result": {}})

    assert len(received) == 2, "only approval_requested and tool.failed should have triggered notifications"

    notification_manager._channels.remove(channel)
    print("[M5-S2] Notification event wiring OK - approval_requested + tool.failed notify, tool.succeeded does not")
def test_excel_read_tools():
    """M7-S1: create a small real .xlsx, then confirm list/read/search
    all work correctly via call_tool, using streaming read-only mode."""
    import openpyxl
    from pathlib import Path
    from app.core.call_tool import call_tool

    test_file = Path("data") / "_test_excel_m7s1.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Score"])
    ws.append(["Alice", 90])
    ws.append(["Bob", 85])
    wb.save(test_file)

    try:
        list_result = call_tool("excel_list_sheets", {"file_path": str(test_file)})
        assert list_result.success is True
        assert any(s["name"] == "Sheet1" for s in list_result.data)

        read_result = call_tool("excel_read_sheet", {"file_path": str(test_file)})
        assert read_result.success is True
        assert read_result.data["rows"][0] == ["Name", "Score"]
        assert read_result.data["rows"][1] == ["Alice", 90]

        search_result = call_tool("excel_search_in_sheet", {"file_path": str(test_file), "query": "alice"})
        assert search_result.success is True
        assert len(search_result.data["matches"]) == 1
        assert search_result.data["matches"][0]["value"] == "Alice"

        print("[M7-S1] Excel read tools OK - list/read/search all correct")
    finally:
        test_file.unlink(missing_ok=True)


def test_excel_invalid_file():
    """M7-S1: confirm a missing file and a wrong extension both raise
    ToolExecutionError cleanly (call_tool wraps any exception raised
    DURING tool execution as ToolExecutionError - only Pydantic's own
    pre-execution input-schema validation raises bare ValidationError)."""
    from app.core.call_tool import call_tool
    from app.core.exceptions import ToolExecutionError

    try:
        call_tool("excel_list_sheets", {"file_path": "does_not_exist.xlsx"})
        assert False, "should have raised for a missing file"
    except ToolExecutionError:
        pass

    try:
        call_tool("excel_list_sheets", {"file_path": "main.py"})
        assert False, "should have raised for a non-Excel file"
    except ToolExecutionError:
        pass

    print("[M7-S1] Excel invalid-file handling OK")

def test_excel_write_tools():
    """M7-S2: confirm write_cell/append_row/create_sheet all persist to
    disk correctly and are approval-gated."""
    import openpyxl
    from pathlib import Path
    from app.core.call_tool import call_tool
    from app.core.approval import AutoApproveHandler
    from app.core.exceptions import ToolExecutionError

    test_file = Path("data") / "_test_excel_m7s2.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.append(["Name", "Score"])
    wb.save(test_file)

    try:
        try:
            call_tool("excel_write_cell", {"file_path": str(test_file), "cell": "C1", "value": "x"})
            assert False, "should require approval"
        except ToolExecutionError:
            pass

        call_tool(
            "excel_write_cell",
            {"file_path": str(test_file), "cell": "C1", "value": "Grade"},
            approval_handler=AutoApproveHandler(),
        )
        call_tool(
            "excel_append_row",
            {"file_path": str(test_file), "row_values": ["Carol", 95]},
            approval_handler=AutoApproveHandler(),
        )
        call_tool(
            "excel_create_sheet",
            {"file_path": str(test_file), "sheet_name": "Notes"},
            approval_handler=AutoApproveHandler(),
        )

        wb_check = openpyxl.load_workbook(test_file)
        assert wb_check.active["C1"].value == "Grade"
        assert list(wb_check.active.iter_rows(values_only=True))[-1][:2] == ("Carol", 95)
        assert "Notes" in wb_check.sheetnames

        print("[M7-S2] Excel write tools OK - write_cell/append_row/create_sheet all persisted correctly")
    finally:
        test_file.unlink(missing_ok=True)

def test_pdf_read_tools():
    """M8-S1: create a small real PDF, then confirm metadata/extract/search
    all work correctly via call_tool."""
    from pathlib import Path
    from pypdf import PdfWriter
    from reportlab.pdfgen import canvas
    from app.core.call_tool import call_tool

    test_file = Path("data") / "_test_pdf_m8s1.pdf"

    # Build a real 2-page PDF with actual extractable text.
    c = canvas.Canvas(str(test_file))
    c.drawString(100, 750, "Page one contains the word banana somewhere.")
    c.showPage()
    c.drawString(100, 750, "Page two talks about apples instead.")
    c.showPage()
    c.save()

    try:
        meta_result = call_tool("pdf_get_metadata", {"file_path": str(test_file)})
        assert meta_result.success is True
        assert meta_result.data["page_count"] == 2

        extract_result = call_tool("pdf_extract_text", {"file_path": str(test_file), "start_page": 1, "end_page": 2})
        assert extract_result.success is True
        assert "banana" in extract_result.data["pages"][0]["text"].lower()
        assert "apples" in extract_result.data["pages"][1]["text"].lower()

        search_result = call_tool("pdf_search_text", {"file_path": str(test_file), "query": "banana"})
        assert search_result.success is True
        assert search_result.data["matches"][0]["page"] == 1

        print("[M8-S1] PDF read tools OK - metadata/extract/search all correct")
    finally:
        test_file.unlink(missing_ok=True)

def main():
    print("[M1-S1] Scaffold check starting...")
    print("[M1-S1] Scaffold check complete.")

    configure_logging()

    print("\n[M1-S2] Config + Logging checks starting...")
    test_config_loads()
    test_logging_writes()
    print("[M1-S2] Config + Logging checks complete.")

    print("\n[M1-S3] Database checks starting...")
    test_database_init()
    print("[M1-S3] Database checks complete.")

    print("\n[M1-S4] Plugin registry checks starting...")
    test_plugin_discovery()
    print("[M1-S4] Plugin registry checks complete.")

    print("\n[M2-S1] call_tool pipeline checks starting...")
    test_call_tool_success()
    test_call_tool_unknown_tool()
    test_call_tool_invalid_input()
    test_call_tool_audit_logged()
    print("[M2-S1] call_tool pipeline checks complete.")

    print("\n[M2-S2] Approval gating checks starting...")
    test_call_tool_read_needs_no_approval()
    test_call_tool_modify_denied_without_handler()
    test_call_tool_modify_approved()
    test_call_tool_modify_denied()
    test_call_tool_batch_approval_per_run()
    print("[M2-S2] Approval gating checks complete.")

    print("\n[M3-S1] Event bus checks starting...")
    test_event_bus_basic_pubsub()
    test_event_bus_subscriber_exception_does_not_propagate()
    test_call_tool_publishes_events()
    print("[M3-S1] Event bus checks complete.")

    print("\n[M3-S2] Approval event checks starting...")
    test_approval_events()
    print("[M3-S2] Approval event checks complete.")

    print("\n[M4-S1] Knowledge base checks starting...")
    test_knowledge_base_crud()
    test_knowledge_base_invalid_content_type()
    print("[M4-S1] Knowledge base checks complete.")

    print("\n[M4-S2] Knowledge base tools checks starting...")
    test_knowledge_tools()
    print("[M4-S2] Knowledge base tools checks complete.")

    print("\n[M5-S1] Notification framework checks starting...")
    test_notification_channel_console()
    test_notification_manager_broadcasts_to_all_channels()
    print("[M5-S1] Notification framework checks complete.")

    print("\n[M5-S2] Notification event wiring checks starting...")
    test_notification_event_wiring()
    print("[M5-S2] Notification event wiring checks complete.")

    print("\n[M7-S1] Excel read tools checks starting...")
    test_excel_read_tools()
    test_excel_invalid_file()
    print("[M7-S1] Excel read tools checks complete.")

    print("\n[M7-S2] Excel write tools checks starting...")
    test_excel_write_tools()
    print("[M7-S2] Excel write tools checks complete.")

    print("\n[M8-S1] PDF read tools checks starting...")
    test_pdf_read_tools()
    print("[M8-S1] PDF read tools checks complete.")

if __name__ == "__main__":
    main()