"""
Excel WRITE tools - write_cell, append_row, create_sheet. All MODIFY
permission (approval-gated via call_tool).

Uses a NORMAL (non-read-only) workbook load, unlike read_tools.py -
read_only=True workbooks cannot be saved in openpyxl, so writing
inherently requires the full workbook in memory. This is a deliberate,
documented exception to the project's streaming-read default (Section
2's 8GB RAM discipline), not an oversight - openpyxl has no streaming
write-to-existing-file mode. Fine for realistic personal spreadsheets;
worth knowing if this is ever pointed at a genuinely enormous file.

Saves go through a temp-file-then-atomic-replace pattern (os.replace,
atomic on both Windows and POSIX) rather than saving in place, so a
failure mid-write can never leave the original file half-written/corrupt.
"""

import os
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.core.exceptions import ValidationError
from app.models.tool_result import ToolResult
from app.registry.tool_contract import PermissionLevel, tool
from plugins.excel._shared import validate_excel_path


def _open_workbook_for_write(file_path: str):
    """Normal (non-read-only) load - required since we need to save afterward."""
    path = validate_excel_path(file_path)
    try:
        return openpyxl.load_workbook(str(path), data_only=False)
    except Exception as e:
        raise ValidationError(f"Failed to open Excel file '{file_path}' for writing: {e}") from e


def _save_atomically(wb, target_path: Path) -> None:
    """
    Saves to a temp file in the SAME directory as target_path (so
    os.replace stays within the same filesystem/volume - required for
    atomicity on some platforms), then atomically replaces the original.
    The original is only ever fully replaced or left completely
    untouched - never partially written.
    """
    fd, temp_path_str = tempfile.mkstemp(suffix=".xlsx", dir=str(target_path.parent))
    os.close(fd)  # we only needed mkstemp to reserve a unique filename
    temp_path = Path(temp_path_str)
    try:
        wb.save(str(temp_path))
        os.replace(str(temp_path), str(target_path))  # atomic on Windows and POSIX
    except Exception:
        temp_path.unlink(missing_ok=True)  # clean up the temp file if save/replace failed
        raise


class WriteCellInput(BaseModel):
    file_path: str = Field(description="Path to the .xlsx/.xlsm file")
    cell: str = Field(description="Cell address, e.g. 'B3'")
    value: object = Field(description="Value to write into the cell")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name; defaults to the active sheet")


@tool(
    name="excel_write_cell",
    description="Write a value into a specific cell of an Excel sheet, saving the file atomically.",
    permission=PermissionLevel.MODIFY,
    input_schema=WriteCellInput,
)
def excel_write_cell(input_data: WriteCellInput) -> ToolResult:
    path = validate_excel_path(input_data.file_path)
    wb = _open_workbook_for_write(input_data.file_path)
    try:
        sheet = wb[input_data.sheet_name] if input_data.sheet_name else wb.active
        sheet[input_data.cell] = input_data.value
        _save_atomically(wb, path)
        return ToolResult(success=True, data={
            "sheet_name": sheet.title,
            "cell": input_data.cell,
            "value": input_data.value,
        })
    finally:
        wb.close()


class AppendRowInput(BaseModel):
    file_path: str = Field(description="Path to the .xlsx/.xlsm file")
    row_values: list = Field(description="Values for the new row, in column order")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name; defaults to the active sheet")


@tool(
    name="excel_append_row",
    description="Append a new row of values to the end of an Excel sheet, saving the file atomically.",
    permission=PermissionLevel.MODIFY,
    input_schema=AppendRowInput,
)
def excel_append_row(input_data: AppendRowInput) -> ToolResult:
    path = validate_excel_path(input_data.file_path)
    wb = _open_workbook_for_write(input_data.file_path)
    try:
        sheet = wb[input_data.sheet_name] if input_data.sheet_name else wb.active
        sheet.append(input_data.row_values)
        new_row_number = sheet.max_row
        _save_atomically(wb, path)
        return ToolResult(success=True, data={
            "sheet_name": sheet.title,
            "row_number": new_row_number,
            "values": input_data.row_values,
        })
    finally:
        wb.close()


class CreateSheetInput(BaseModel):
    file_path: str = Field(description="Path to the .xlsx/.xlsm file")
    sheet_name: str = Field(description="Name for the new sheet")


@tool(
    name="excel_create_sheet",
    description="Create a new, empty sheet in an Excel workbook. Refuses to overwrite an existing sheet name.",
    permission=PermissionLevel.MODIFY,
    input_schema=CreateSheetInput,
)
def excel_create_sheet(input_data: CreateSheetInput) -> ToolResult:
    path = validate_excel_path(input_data.file_path)
    wb = _open_workbook_for_write(input_data.file_path)
    try:
        if input_data.sheet_name in wb.sheetnames:
            raise ValidationError(
                f"Sheet '{input_data.sheet_name}' already exists - refusing to overwrite it"
            )
        wb.create_sheet(input_data.sheet_name)
        _save_atomically(wb, path)
        return ToolResult(success=True, data={"sheet_name": input_data.sheet_name, "created": True})
    finally:
        wb.close()