"""
Excel READ tools - list sheets, read a range, search for values.
All READ permission (no approval gate). Write/modify tools are a
separate module (write_tools.py, M7-S2) since they're higher-risk and
use a different (non-read-only) workbook-opening mode.
"""

from typing import Optional

from pydantic import BaseModel, Field

from app.models.tool_result import ToolResult
from app.registry.tool_contract import PermissionLevel, tool
from plugins.excel._shared import open_workbook_readonly
from openpyxl.utils import get_column_letter

class ListSheetsInput(BaseModel):
    file_path: str = Field(description="Path to the .xlsx/.xlsm file")

@tool(
    name="excel_list_sheets",
    description="List all sheet names in an Excel workbook, along with each sheet's dimensions.",
    permission=PermissionLevel.READ,
    input_schema=ListSheetsInput,
)
def excel_list_sheets(input_data: ListSheetsInput) -> ToolResult:
    wb = open_workbook_readonly(input_data.file_path)
    try:
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            # .dimensions isn't available on ReadOnlyWorksheet (read_only=True
            # mode doesn't eagerly parse the full used range) - compute the
            # same "A1:F10"-style string manually from max_row/max_column instead.
            if ws.max_row and ws.max_column:
                dimensions = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            else:
                dimensions = "empty"
            sheets.append({"name": name, "dimensions": dimensions})
        return ToolResult(success=True, data=sheets)
    finally:
        wb.close()  # important in read_only mode - releases the underlying file handle


class ReadRangeInput(BaseModel):
    file_path: str = Field(description="Path to the .xlsx/.xlsm file")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name; defaults to the active sheet")
    max_rows: int = Field(default=200, description="Maximum number of rows to read (streaming cap)")


@tool(
    name="excel_read_sheet",
    description=(
        "Read rows from an Excel sheet as a list of row lists, streamed "
        "efficiently (read_only mode) rather than loading the whole file "
        "into memory. Capped at max_rows to avoid returning an enormous result."
    ),
    permission=PermissionLevel.READ,
    input_schema=ReadRangeInput,
)
def excel_read_sheet(input_data: ReadRangeInput) -> ToolResult:
    wb = open_workbook_readonly(input_data.file_path)
    try:
        sheet = wb[input_data.sheet_name] if input_data.sheet_name else wb.active

        rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= input_data.max_rows:
                break
            rows.append(list(row))

        return ToolResult(success=True, data={
            "sheet_name": sheet.title,
            "row_count_returned": len(rows),
            "rows": rows,
        })
    finally:
        wb.close()


class SearchInSheetInput(BaseModel):
    file_path: str = Field(description="Path to the .xlsx/.xlsm file")
    query: str = Field(description="Text to search for within cell values (case-insensitive substring match)")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name; defaults to the active sheet")
    max_results: int = Field(default=50, description="Stop after this many matches")


@tool(
    name="excel_search_in_sheet",
    description="Search a sheet for cells whose value contains the given text, returning matching cell addresses and values.",
    permission=PermissionLevel.READ,
    input_schema=SearchInSheetInput,
)
def excel_search_in_sheet(input_data: SearchInSheetInput) -> ToolResult:
    wb = open_workbook_readonly(input_data.file_path)
    try:
        sheet = wb[input_data.sheet_name] if input_data.sheet_name else wb.active
        query_lower = input_data.query.lower()
        matches = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and query_lower in str(cell.value).lower():
                    matches.append({"cell": cell.coordinate, "value": cell.value})
                    if len(matches) >= input_data.max_results:
                        return ToolResult(success=True, data={
                            "sheet_name": sheet.title,
                            "matches": matches,
                            "truncated": True,
                        })

        return ToolResult(success=True, data={
            "sheet_name": sheet.title,
            "matches": matches,
            "truncated": False,
        })
    finally:
        wb.close()