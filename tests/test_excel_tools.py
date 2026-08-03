"""
Regression suite for Excel tools (M7) - read tools (streaming read-only)
and write tools (atomic save), using real temp .xlsx files, no mocking
of openpyxl itself since correctness here depends on real file I/O
behavior.
"""

import openpyxl
import pytest

from app.core.approval import AutoApproveHandler
from app.core.call_tool import call_tool
from app.core.exceptions import ToolExecutionError


@pytest.fixture
def sample_workbook(tmp_path):
    """Creates a real small .xlsx in a temp dir, returns its path as a string."""
    file_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score"])
    ws.append(["Alice", 90])
    ws.append(["Bob", 85])
    wb.save(file_path)
    return str(file_path)


class TestExcelReadTools:
    def test_list_sheets(self, sample_workbook, isolated_db):
        result = call_tool("excel_list_sheets", {"file_path": sample_workbook})
        assert result.success is True
        assert result.data[0]["name"] == "Data"

    def test_read_sheet(self, sample_workbook, isolated_db):
        result = call_tool("excel_read_sheet", {"file_path": sample_workbook})
        assert result.success is True
        assert result.data["rows"][1] == ["Alice", 90]

    def test_search_in_sheet(self, sample_workbook, isolated_db):
        result = call_tool("excel_search_in_sheet", {"file_path": sample_workbook, "query": "bob"})
        assert result.success is True
        assert len(result.data["matches"]) == 1

    def test_missing_file_raises(self, isolated_db):
        with pytest.raises(ToolExecutionError):
            call_tool("excel_list_sheets", {"file_path": "nope.xlsx"})


class TestExcelWriteTools:
    def test_write_cell_requires_approval(self, sample_workbook, isolated_db):
        with pytest.raises(ToolExecutionError):
            call_tool("excel_write_cell", {"file_path": sample_workbook, "cell": "C1", "value": "new"})

    def test_write_cell_persists(self, sample_workbook, isolated_db):
        call_tool(
            "excel_write_cell",
            {"file_path": sample_workbook, "cell": "C1", "value": "Grade"},
            approval_handler=AutoApproveHandler(),
        )
        # Re-open independently to confirm the write actually persisted to disk.
        wb = openpyxl.load_workbook(sample_workbook)
        assert wb.active["C1"].value == "Grade"

    def test_append_row_persists(self, sample_workbook, isolated_db):
        call_tool(
            "excel_append_row",
            {"file_path": sample_workbook, "row_values": ["Carol", 95]},
            approval_handler=AutoApproveHandler(),
        )
        wb = openpyxl.load_workbook(sample_workbook)
        rows = list(wb.active.iter_rows(values_only=True))
        assert rows[-1] == ("Carol", 95)

    def test_create_sheet_persists(self, sample_workbook, isolated_db):
        call_tool(
            "excel_create_sheet",
            {"file_path": sample_workbook, "sheet_name": "Notes"},
            approval_handler=AutoApproveHandler(),
        )
        wb = openpyxl.load_workbook(sample_workbook)
        assert "Notes" in wb.sheetnames

    def test_create_sheet_refuses_duplicate(self, sample_workbook, isolated_db):
        with pytest.raises(ToolExecutionError):
            call_tool(
                "excel_create_sheet",
                {"file_path": sample_workbook, "sheet_name": "Data"},  # already exists
                approval_handler=AutoApproveHandler(),
            )